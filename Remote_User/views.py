from django.db.models import Count
from django.db.models import Q
from django.shortcuts import render, redirect, get_object_or_404
import datetime
import openpyxl

import pandas as pd
from sklearn.metrics import accuracy_score, confusion_matrix, classification_report
from sklearn.ensemble import VotingClassifier
import pandas as pd
import numpy as np

# Create your views here.
from Remote_User.models import ClientRegister_Model,Tweet_Message,Tweet_Type_Prediction,detection_ratio,detection_accuracy

def login(request):


    if request.method == "POST" and 'submit1' in request.POST:

        username = request.POST.get('username')
        password = request.POST.get('password')
        try:
            enter = ClientRegister_Model.objects.get(username=username,password=password)
            request.session["userid"] = enter.id

            return redirect('Add_DataSet_Details')
        except:
            pass

    return render(request,'RUser/login.html')

def Add_DataSet_Details(request):
    if "GET" == request.method:
        return render(request, 'RUser/Add_DataSet_Details.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        #print(sheets)
        # getting a particular sheet
        worksheet = wb["Sheet1"]
        #print(worksheet)
        # getting active sheet
        active_sheet = wb.active
        #print(active_sheet)
        # reading a cell
        #print(worksheet["A1"].value)
        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                #print(cell.value)
            excel_data.append(row_data)
            Tweet_Message.objects.all().delete()
            Tweet_Type_Prediction.objects.all().delete()
    for r in range(1, active_sheet.max_row + 1):
        Tweet_Message.objects.create(
        idno=active_sheet.cell(r, 1).value,
        Tweet=active_sheet.cell(r, 2).value,
        following=active_sheet.cell(r, 3).value,
        followers=active_sheet.cell(r, 4).value,
        actions=active_sheet.cell(r, 5).value,
        is_retweet=active_sheet.cell(r, 6).value,
        location=active_sheet.cell(r, 7).value
        )

    return render(request, 'RUser/Add_DataSet_Details.html', {"excel_data": excel_data})


def Register1(request):

    if request.method == "POST":
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        phoneno = request.POST.get('phoneno')
        country = request.POST.get('country')
        state = request.POST.get('state')
        city = request.POST.get('city')
        ClientRegister_Model.objects.create(username=username, email=email, password=password, phoneno=phoneno,
                                            country=country, state=state, city=city)

        return render(request, 'RUser/Register1.html')
    else:
        return render(request,'RUser/Register1.html')

def ViewYourProfile(request):
    userid = request.session['userid']
    obj = ClientRegister_Model.objects.get(id= userid)
    return render(request,'RUser/ViewYourProfile.html',{'object':obj})


def Search_DataSets(request):
    if request.method == "POST":
        kword = request.POST.get('keyword')
        if request.method == "POST":
            kword = request.POST.get('keyword')

        tweet_df = pd.read_csv('data_train.csv')
        tweet_df
        tweet_df.columns
        tweet_df.drop(['Id', 'following', 'followers', 'actions', 'is_retweet', 'location'], axis=1, inplace=True)
        tweet_df['Type'] = tweet_df['Type'].map({'Quality': 0, 'Bot': 1})
        from sklearn.feature_extraction.text import CountVectorizer
        cv = CountVectorizer()
        x = tweet_df['Tweet']
        y = tweet_df['Type']
        x = cv.fit_transform(x)

        from sklearn.model_selection import train_test_split

        x_train, x_test, y_train, y_test = train_test_split(x, y, test_size=0.2)
        x_train.shape

        models = []
        from sklearn.neighbors import KNeighborsClassifier
        model = KNeighborsClassifier(n_neighbors=3, weights='distance')
        models.append(('KNeighborsClassifier', model))

        from sklearn import svm
        lin_clf = svm.LinearSVC()
        models.append(('svm', lin_clf))

        from sklearn.linear_model import LogisticRegression
        reg = LogisticRegression(random_state=0, solver='lbfgs').fit(x_train, y_train)
        models.append(('logistic', reg))

        from sklearn.naive_bayes import MultinomialNB
        NB = MultinomialNB()
        NB.fit(x_train, y_train)

        classifier = VotingClassifier(models)
        classifier.fit(x_train, y_train)
        y_pred = classifier.predict(x_test)

        tweet_data = [kword]
        vector1 = cv.transform(tweet_data).toarray()
        predict_text = classifier.predict(vector1)
        if predict_text == 1:
             val = 'Bot'
        else:
            val = 'Quality'

        return render(request, 'RUser/Search_DataSets.html',{'objs': val})
    return render(request, 'RUser/Search_DataSets.html')



